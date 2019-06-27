#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# @Time    : 4/25/2019 11:48 AM
# @Author  : Owen_study
# @Email   : owen_study@126.com
# @File    : getloadinglogsummary.py
# @Software: PyCharm Community Edition
# ===============================================
'''
检查sqlloader 执行后多个日志文件的日志信息，检查后生成各个文件的加载情况，保存到excel中
文件放在sqlldr上一级目录，以方便在加载后调用
'''
import os,openpyxl
from openpyxl.workbook import Workbook

# 检查一个文件的内容，并输出文件的关键内容
# 输出的结果形式 filename|loadtotal|errortotal|errormessage
def checklog(filename):
    logfile=open(filename, 'r')
    summary='{filename}|{loadtotal}|{errortotal}|{errormessage}'
    loadtotal=0
    errortotal=0
    errormessage = ''
    for eachline in logfile:
        # 查找成功加载的数量
        succpos=eachline.find('Rows successfully loaded')
        if (succpos>0):
            loadtotal =eachline[:succpos].replace(' ','')
        # 查找加载失败的数量
        errorpos=eachline.find('Rows not loaded due to data errors')
        if (errorpos > 0):
            errortotal = eachline[:errorpos].replace(' ', '')
            pass
        # 加载报错信息
        errormsgpos=eachline.find('SQL*Loader-')
        if (errormsgpos >= 0):
            errormessage=errormessage+eachline
            logfile.close()
            break
    # 生成合并的日志信息
    summary=summary.format(filename=os.path.split(filename)[1],loadtotal=loadtotal,errortotal=errortotal,errormessage=errormessage)
    logfile.close()
    return summary
    print(summary)
    pass
# 把文件结果输出到excel中
def outputexcel(rowindex,eachfilesummary,filepath):
    # 把结果生成到一个文件中
    # summyexcel = Workbook()
    # filepath=os.path.dirname(os.path.dirname(__file__))
    # print(filepath)
    # 把文件保存在上一级目录，以方便生成exe文件时执行时放在主目录
    if (os.path.exists(os.path.join(filepath ,'SqlloaderLogSummary.xlsx'))):
        summyexcel=openpyxl.load_workbook(os.path.join(filepath ,'SqlloaderLogSummary.xlsx'), data_only=True)
    else:
        summyexcel = Workbook()
    wssummary = summyexcel.active
    wssummary.title = 'SqlloaderLogSummary'

    # 设置每行的列名以方便查看
    colindex = 'A'
    columntitle = ['FileName','LoadNumber','ErrorNumber','ErrorMessage']
    for column in columntitle:
        cellindex = '{0}{1}'.format(colindex, 1)
        wssummary[cellindex] = column
        colindex = chr(ord(colindex) + 1)
    # 保存内容到指定的行
    # 循环处理传送过来的内容，分散到各个列中
    colindex = 'A'
    for eachvalue in eachfilesummary.split('|'):
        wssummary['{0}{1}'.format(colindex, rowindex)].value = eachvalue
        colindex = chr(ord(colindex) + 1)

    # 保存文件到磁盘
    summyexcel.save( os.path.join(filepath ,'SqlloaderLogSummary.xlsx'))

    pass
#  对指定目录下的文件进行循环处理
# 只查找sqlldr目录log下面的文件
def checkalllogs(filepath):
    # 只查找指定的目录名称
    # filepath=os.path.dirname(os.path.dirname(__file__))
    print(filepath)
    logpath=os.path.join(filepath, 'sqlldr/log')
    if (os.path.isdir(logpath)):
        index=2
        for eachfile in os.listdir(logpath):
            # 处理每一个文件
            summary=checklog(os.path.join(logpath,eachfile))
            # 保存excel文件中
            outputexcel(index,summary,filepath)
            index=index+1
            # print(summary)
        # print('has log folder')
    else:
        print('There is no log folder is found!')
    # 查找日志下面所有的日志文件

    pass


if __name__ == '__main__':
    # outputexcel(2,'xxx|6|0|xxx')
    checkalllogs('D://allLoadVeriScripts')
    # checklog('./sqlldr/log/DM_PARTY.log')
    # checklog('./sqlldr/log/DM_EXTRA_PREM.log')

    pass