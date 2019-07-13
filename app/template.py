#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = 'Owen_Study/owen_study@126.com'

import openpyxl,re, os
from openpyxl import load_workbook
from numpy import nan
from app import public_init_script, configure

from app import common
import numba, numpy as np, pandas as pd

'''从excel加载template'''
class Template(object):
    @numba.jit()
    # 初始化数据文件到object
    def __init__(self,file_name,ignore_strike_row=True):
        self.__file_name=file_name
        self.excel_handler=openpyxl.load_workbook(file_name, data_only=True)
        self.__ignore_strike_row=ignore_strike_row

    # test template cols
    def test__get_mapping_cols(self):
        mapping_columns_sheet=self.__get_mapping_cols_sheet(self.excel_handler)
        mapping_columns_list=self.get_mapping_cols(mapping_columns_sheet)



    # 取得mapping columns这个sheet
    def __get_mapping_cols_sheet(self,excel_handler):
        sheet_names = excel_handler.sheetnames
        mapping_columns_sheet=None
        for sheet in sheet_names:
            if sheet=='MappingColumns':
                mapping_columns_sheet=excel_handler.get_sheet_by_name(sheet)
                break
        return mapping_columns_sheet

    @numba.jit()
    # 读取excel columns中的字段信息，放到json对象中
    def get_mapping_cols(self):
        mapping_columns_sheet=self.__get_mapping_cols_sheet(self.excel_handler)

        # Template Table Name	Column Name	Data Type	Length	Nullable	Key	Short Description	Descirption of Data Migration	Default Value	Reference Table
        #所有的列名
        column_title_name=('tableName','columnName','dataType','length','nullable','primaryKey','descShort','descDM','defaultValue','referTable')
        # 循环所有的行
        mapping_rows=[]
        for row in mapping_columns_sheet.rows:
            # 循环所有 的列
            cell_value={}
            for cell in row:
                #忽略已经加了删除线的行,只检查前两列是否有删除线，则说明本行是删除的，其它单元格的删除线不识别
                if cell.font.strikethrough and cell.col_idx<=2 and self.__ignore_strike_row:
                    break
                # 排除掉第一行的标题
                if cell.col_idx<=len(column_title_name) and cell.row>=2:
                    title_name=column_title_name[cell.col_idx-1]
                    cell_value[title_name]=cell.value
                    # debug
                    if cell.value == 'COUNTRY':
                        pass
                        # print('stop')
                    #print(cell_value)
            if len(cell_value)>0:
                # 去除文件名的路径信息
                module_name = os.path.split(self.__file_name)[1]
                cell_value['moduleName']=module_name
                cellobj=common.JSONObject(cell_value)
                mapping_rows.append(cellobj)
        return mapping_rows

class DCDocConfigExcel(object):
    # 初始化数据文件到object
    def __init__(self,file_name,ignore_strike_row=True):
        self.__file_name=file_name
        self.excel_handler=openpyxl.load_workbook(file_name, data_only=True)
        self.__ignore_strike_row=ignore_strike_row
    # 取得excel配置的数据信息
    def get_config_data(self):
        docconfig = self.excel_handler.worksheets[0]
        #所有的列名,如果新增加列则直接增加列名,返回的内容只包括列表所在的数据内容
        column_title_name=('phase','sn','majorinfo','output','remark','referdoc')
        # 循环所有的行
        config_rows=[]
        for row in docconfig.rows:
            # 循环所有 的列
            cell_value={}
            for cell in row:
                #忽略已经加了删除线的行,只检查前两列是否有删除线，则说明本行是删除的，其它单元格的删除线不识别
                if cell.font.strikethrough and cell.col_idx<=2 and self.__ignore_strike_row:
                    break
                # 排除掉第一行的标题
                if cell.col_idx<=len(column_title_name) and cell.row>=2:
                    title_name=column_title_name[cell.col_idx-1]
                    cell_value[title_name]=cell.value
            if len(cell_value) > 0:
                # cellobj = common.JSONObject(cell_value)
                # config_rows.append(cellobj)
                config_rows.append(cell_value)

        # print(config_rows[0].majorinfo)
        for config in config_rows:
            for value in config:
                print(config[value])
        return config_rows
        pass
# 读取映射文档，统计映射信息, 2018.11.8
class DCMappingExcel(object):
    # 初始化数据文件到object
    def __init__(self,file_name,ignore_strike_row=True):
        self.__file_name=file_name
        self.excel_handler=openpyxl.load_workbook(file_name, data_only=True)
        self.__ignore_strike_row=ignore_strike_row
    # 取得excel配置的数据信息
    def get_config_data(self):
        docconfig = self.excel_handler.worksheets[2]
        #所有的列名,如果新增加列则直接增加列名,返回的内容只包括列表所在的数据内容
        column_title_name=('tableName','columnName','dataType','length','nullable','primaryKey','descShort','descDM','defaultValue','referTable', \
                           'sourcetable','sourcefield', 'mappingrule', 'comment', 'mappingdone', 'developeddone')
        # 循环所有的行
        config_rows=[]
        for row in docconfig.rows:
            # 循环所有 的列
            cell_value={}
            for cell in row:
                #忽略已经加了删除线的行,只检查前两列是否有删除线，则说明本行是删除的，其它单元格的删除线不识别
                if cell.font.strikethrough and cell.col_idx<=2 and self.__ignore_strike_row:
                    break
                # 排除掉第一行的标题
                if cell.col_idx<=len(column_title_name) and cell.row>=2:
                    title_name=column_title_name[cell.col_idx-1]
                    cell_value[title_name]=cell.value
            if len(cell_value) > 0:
                # cellobj = common.JSONObject(cell_value)
                # config_rows.append(cellobj)
                config_rows.append(cell_value)

        # print(config_rows[0].majorinfo)
        for config in config_rows:
            for value in config:
                pass
                # print(config[value])
        return config_rows
        pass

class DCReport(object):
    # 初始化数据文件到object
    def __init__(self,file_name,ignore_strike_row=True):
        self.__file_name=file_name
        self.excel_handler=openpyxl.load_workbook(file_name, data_only=True)
        self.__ignore_strike_row=ignore_strike_row
        # 报表内容的列名
        self.column_title_name=('BRR_CODE','Column1','Column2','Column3','Column4')
        # BRR定义的列的内容
        self.BRR_COLUMN_LIST = ('BRR_Status','SN','Module','BRR_CODE','BRR_Desc','BRR_Column1','BRR_Column2','BRR_Column3','Cnt_Column1','Comments')
        # 固定的sheet顺序
        self.REPORT_BRR_LIST_INDX = 2
        self.REPORT_SUMMARY_INDX = 4
        self.REPORT_DIFF_INDX = 5
        self.TARGET_SHEET_INDX = 6
        self.SOURCE_SHEET_INDX = 7


    # 处理源和目标的报表数据
    def process_report_data(self):
        # 目标库的报表结果
        doctarget = self.excel_handler.worksheets[self.TARGET_SHEET_INDX-1]
        # 源系统中的报表结果
        docsource = self.excel_handler.worksheets[self.SOURCE_SHEET_INDX-1]
        # 获得源和目标的报表数据
        sourcereport = self.get_report_data(docsource,self.column_title_name)
        # 把每个报表的最后一列统计放到一个新列中以方便比较
        updatesourcereport = self.update_report_date(sourcereport)
        # print(updatesourcereport)
        # 目标表的数据获取和更新整理
        targetreport = self.get_report_data(doctarget,self.column_title_name)
        updatetargetreport = self.update_report_date(targetreport)
        # 比较源和目标报表的差异
        mergereport = pd.merge(updatetargetreport,updatesourcereport,how='outer' ,on=self.column_title_name)
        # 对统计列进行重命名以方便理解
        mergereport.rename(columns = {'RESULT_x':'ebao_values','RESULT_y':'legacy_values'}, inplace=True)
        # 每一行增加一列来判断是不是通过或者不通过
        # mergereport['passindi'] = mergereport
        # print(mergereport.columns)
        # 两个报表之间的差异
        diffreport = mergereport[mergereport['ebao_values']!=mergereport['legacy_values']].sort_values(by = ['BRR_CODE','Column1']).copy()

        # 把报表不同的内容输出到reportdiff
        self.output_diff_report(diffreport)

        # 输出报表的summary
        self.generate_report_summary(mergereport)

        # 保存更新的结果
        self.excel_handler.save(self.__file_name)
    # 把报表中不同的部分追加到excel明细文件中
    def output_diff_report(self, diffreport):
        # 定义的sheet顺序为第4个是差异，如果有变动就会覆盖
        sheetname = self.excel_handler.sheetnames[self.REPORT_DIFF_INDX-1]
        # 清除已经存在的内容
        self.clear_sheet(self.excel_handler,sheetname)
        # print(sheetname)
        # 把差异结果输出到excel
        exceldiff = self.excel_handler.worksheets[self.REPORT_DIFF_INDX-1]
        # 差异的标题内容
        difftitle = ['Module','BRR_CODE','BRR_Desc','Comments','Column1','Column2','Column3','eBaoValues','LegacyValues']
        exceldiff.append(difftitle)

        # 获取brr list
        doctbrrlist = self.excel_handler.worksheets[self.REPORT_BRR_LIST_INDX-1]
        # 只加载BRR列表定义中的前几列
        brrlistdata = self.get_report_data(doctbrrlist,self.BRR_COLUMN_LIST)
        # 获取BRR status=Y的列表
        validatebrrlist = brrlistdata[brrlistdata['BRR_Status'] == 'Y']
        # 和BRR定义表关联使生成的报表可读性更强
        # 把BRR_CODE的描述信息合并进来以方便报表生成
        differreportwithdesc = pd.merge(diffreport,validatebrrlist,how='inner',on='BRR_CODE')
        print(differreportwithdesc.columns)
        # 只显示需要的列
        showdiffreport = differreportwithdesc[['Module','BRR_CODE','BRR_Desc','Comments','Column1','Column2','Column3','ebao_values','legacy_values']]
        self.output_excel(showdiffreport,difftitle,self.REPORT_DIFF_INDX)

        # diffreport.fillna(value=nan, inplace= True)
        # 要转换成矩阵来输出到excel中
        # diffdata = diffreport.as_matrix()
        # for rowdata in diffdata:
        #     # print(str(rowdata))
        #     # 必须转为list才能追加到excel列表中
        #     exceldiff.append(list(rowdata))
        # 保存更新的结果，不能重复保存，会出现问题
        # self.excel_handler.save(self.__file_name)
    # 把报表中不同的部分追加到excel明细文件中
    def output_excel(self, pdreport, exceltitle, sheetindex):
        # 定义的sheet顺序为第4个是差异，如果有变动就会覆盖
        sheetname = self.excel_handler.sheetnames[sheetindex-1]
        # 清除已经存在的内容
        self.clear_sheet(self.excel_handler,sheetname)
        # print(sheetname)
        # 把差异结果输出到excel
        excelsheet = self.excel_handler.worksheets[sheetindex-1]
        # 差异的标题内容
        excelsheet.append(exceltitle)

        # diffreport.fillna(value=nan, inplace= True)
        # 要转换成矩阵来输出到excel中
        reportdata = pdreport.as_matrix()
        for rowdata in reportdata:
            # 必须转为list才能追加到excel列表中
            excelsheet.append(list(rowdata))
        # 保存更新的结果
        # self.excel_handler.save(self.__file_name)

    # 生成报表的summary内容
    def generate_report_summary(self,mergereport):
        # 获取brr list
        doctbrrlist = self.excel_handler.worksheets[self.REPORT_BRR_LIST_INDX-1]
        # 只加载BRR列表定义中的前几列
        brrlistdata = self.get_report_data(doctbrrlist,self.BRR_COLUMN_LIST)
        # 获取BRR status=Y的列表
        validatebrrlist = brrlistdata[brrlistdata['BRR_Status'] == 'Y']
        # 增加一列来判断
        # 对合并的结果进行分组，统计每个报表的总数和通过的数量，以目标表的数据为准进行统计，因为源报表给的数据可能和目标表的数据完全对不上
        # brrgroup = mergereport.groupby('BRR_CODE').size().reset_index(name='totalcount')
        # 对每个报表进行分组统计
        brrgroup = mergereport[mergereport['ebao_values'].notnull()].groupby('BRR_CODE').size().reset_index(name='totalcount')
        # 比较结果不能过的数据，只按目标来统计，对于源系统存在目标不存在的记录不进行统计
        brrunmatchgroup = mergereport[(mergereport['ebao_values']!=mergereport['legacy_values']) & (mergereport['ebao_values'].notnull())].groupby('BRR_CODE').size().reset_index(name='diffcount')
        # 合并统计信息，每个报表的总数量和未通过的数量
        groupsummary = pd.merge(brrgroup,brrunmatchgroup,how='outer' ,on='BRR_CODE')
        # 对于没有值的diff count设置为0
        groupsummary['diffcount'].fillna(value=0, inplace=True)
        # 计算通过率
        groupsummary['passrate']= round((groupsummary['totalcount'] - groupsummary['diffcount'])/groupsummary['totalcount'],3)
        # 把通过率的信息汇总到summar sheet
        summarytitle = ['Module','BRR_CODE','BRR_Desc','Comments','TotalRecordsCount','DiffRecordsCount','PassRate','ReportComments']
        # 和BRR定义表关联使生成的报表可读性更强
        # 把BRR_CODE的描述信息合并进来以方便报表生成
        groupsummary = pd.merge(groupsummary,validatebrrlist,how='inner',on='BRR_CODE')
        print(groupsummary.columns)
        # 只显示需要的列
        showsummary = groupsummary[['Module','BRR_CODE','BRR_Desc','Comments','totalcount','diffcount','passrate']]
        self.output_excel(showsummary,summarytitle,self.REPORT_SUMMARY_INDX)

        pass
    # 清除sheet的内容
    def clear_sheet(self, workbook, sheetname):
        # index of [sheet_name] sheet
        idx = workbook.sheetnames.index(sheetname)
        # remove [sheet_name]
        workbook.remove(workbook.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        workbook.create_sheet(sheetname, idx)
        pass
    # 对报表结果进行预处理，使最后一列单独列出来，以方便后面的比较
    def update_report_date(self, reportdata):
        brr_list = list(reportdata['BRR_CODE'].unique())
        newreport = reportdata.head(0)
        for brr_id in brr_list:
            # 对每一个报表的数据进行处理，把结果放在最后一个新的列里面
            eachbrrdata= reportdata[reportdata['BRR_CODE']==brr_id].copy()
            # 对每一个报表中的列进行循环处理，找到最后一个有值的列，这个列作为统计列
            for eachcolumn in self.column_title_name:
                # 找到最后一个全为空的列，则前面一个列为统计列
                if eachbrrdata[eachcolumn].count() == 0:
                    # 把结果列放在一个新增加的列中以方便后面的对比
                    eachbrrdata.loc[:,'RESULT'] = eachbrrdata[lastcolumn]
                    # 清空原来的统计列信息
                    eachbrrdata.loc[:,lastcolumn] = None
                    # 合并后新的源系统值
                    newreport = pd.concat([newreport,eachbrrdata])
                    break
                # 保存当前列名信息
                lastcolumn = eachcolumn
            pass
        # print(newreport)
        return newreport

    # 取得excel中报表的数据
    def get_report_data(self,reportdata,column_title_name):
        #所有的列名,如果新增加列则直接增加列名,返回的内容只包括列表所在的数据内容
        # 循环所有的行
        report_rows=[]
        data = []
        for row in reportdata.rows:
            # 循环所有 的列
            cell_value={}
            rowdata = []
            for cell in row:
                #忽略已经加了删除线的行,只检查前两列是否有删除线，则说明本行是删除的，其它单元格的删除线不识别
                if cell.font.strikethrough and cell.col_idx<=2 and self.__ignore_strike_row:
                    # 如果某些列没有加删除线的情况需要考虑，已经增加的内容需要删除掉,这个问题折腾了好久才发现报错的原因
                    rowdata = []
                    cell_value={}
                    break
                # 排除掉第一行的标题
                if cell.col_idx<=len(column_title_name) and cell.row>=2:
                    title_name=column_title_name[cell.col_idx-1]
                    cell_value[title_name]=cell.value
                    rowdata.append(cell.value)
            if len(cell_value) > 0:
                report_rows.append(cell_value)
                data.append(rowdata)
        npdata = np.array(data)
        pddata = pd.DataFrame(npdata)
        if len(pddata) !=0:
            pddata.columns = list(column_title_name)
        return pddata

'''2019.7.2 改造成从excel加载数据并生生成SQL语句保存结果到DB中，之后对表中的数据做循环执行'''
class DCVerifySQL(object):
    # 初始化数据文件到object
    def __init__(self,file_name,ignore_strike_row=True):
        self.__file_name=file_name
        self.excel_handler=openpyxl.load_workbook(file_name, data_only=True)
        self.__ignore_strike_row=ignore_strike_row
        self.__sql_file_handler = open(os.path.join(configure.DOWNLOAD_FOLDER,public_init_script.validation_file_name),'w')
    # 生成校验语句的表，保存excel中定义的校验语句
    def __create_verify_log_table(self):
        # 创建表dc_validation的脚本
        sql = public_init_script.init_dc_validation
    # 取得excel校验语句的数据信息，各个sheet中的格式应该保持统一以方便统一生成
    def get_verify_sheet_data(self, sheetindex = 0 ):
        docconfig = self.excel_handler.worksheets[sheetindex]
        #所有的列名,如果新增加列则直接增加列名,返回的内容只包括列表所在的数据内容
        column_title_name=('SN','MODULE','IN_PROJECT','PRIORITY','ERROR_CODE','TABLE_NAME','COLUMN_NAME','DESCRIPTION','SQL_FOR_SOURCE','SQL_FOR_TARGET','REMARK','TEMP_SKIP','RULE_FROM')

        # 循环所有的行
        config_rows=[]
        for row in docconfig.rows:
            # 循环所有 的列
            cell_value={}
            for cell in row:
                #忽略已经加了删除线的行,只检查前两列是否有删除线，则说明本行是删除的，其它单元格的删除线不识别
                if cell.font.strikethrough and cell.col_idx<=2 and self.__ignore_strike_row:
                    break
                # 排除掉第一行的标题
                if cell.col_idx<=len(column_title_name) and cell.row>=2:
                    title_name=column_title_name[cell.col_idx-1]
                    if cell.value is None:
                        cell_value[title_name]= ''
                    else:
                        cell_value[title_name]=self.__sql_adjust(cell.value)
            # 对于 非空的行则增加到变量里面
            if len(cell_value) > 0:
                config_rows.append(cell_value)
        return config_rows
    # 处理sql语句中的特殊字符
    def __sql_adjust(self, sql):
        # 如果sql中有'则做'的处理
        if type(sql)==type('str'):
            newsql = sql.replace("'","''")
        else:
            newsql=sql
        return newsql

        pass
    # 针对已经读取的excel数据，生成insert sql 语句的脚本
    def gen_insert_sql(self):
        all_insert_sql = ''
        # 默认为第二个sheet为校验脚本的sheet, sheet index =1
        validation_rules = self.get_verify_sheet_data(1)
        #所有的列名,如果新增加列则直接增加列名,返回的内容只包括列表所在的数据内容
        column_title_name=('SN','MODULE','IN_PROJECT','PRIORITY','ERROR_CODE','TABLE_NAME','COLUMN_NAME','DESCRIPTION','SQL_FOR_SOURCE','SQL_FOR_TARGET','REMARK','TEMP_SKIP','RULE_FROM')
        for eachvalidation in validation_rules:
            # 对每个列进行处理
            insert_sql = public_init_script.init_insert_dc_validation.format( \
                SN=eachvalidation['SN                '.strip()], \
                MODULE=eachvalidation['MODULE            '.strip()], \
                IN_PROJECT=eachvalidation['IN_PROJECT        '.strip()], \
                PRIORITY=eachvalidation['PRIORITY          '.strip()], \
                ERROR_CODE=eachvalidation['ERROR_CODE        '.strip()], \
                TABLE_NAME=eachvalidation['TABLE_NAME        '.strip()], \
                COLUMN_NAME=eachvalidation['COLUMN_NAME       '.strip()], \
                DESCRIPTION=eachvalidation['DESCRIPTION       '.strip()], \
                SQL_FOR_SOURCE=eachvalidation['SQL_FOR_SOURCE    '.strip()], \
                SQL_FOR_TARGET=eachvalidation['SQL_FOR_TARGET    '.strip()], \
                REMARK=eachvalidation['REMARK            '.strip()], \
                TEMP_SKIP=eachvalidation['TEMP_SKIP         '.strip()], \
                RULE_FROM=eachvalidation['RULE_FROM         '.strip()], \
                )
            all_insert_sql = all_insert_sql + insert_sql
        all_insert_sql = all_insert_sql + 'commit;\n'
        return all_insert_sql
        pass
    # 生成dc_validation表的创建表语句 及insert 语句
    def get_all_sqls(self):
        all_sql = public_init_script.init_dc_validation + self.gen_insert_sql()
        return all_sql
        pass
    # 2019.7.3 生成脚本文件
    def gen_script_file(self):
        filenameonly = self.__file_name.split('/')[-1]
        self.__sql_file_handler.write("spool {0}.log\n".format(public_init_script.validation_file_name.split('.')[0]))
        # 生成insert sql
        insert_sql = self.get_all_sqls()
        self.__sql_file_handler.write(insert_sql)
        self.__sql_file_handler.write("\nspool off")
        self.__sql_file_handler.write("\n quit; ")
        self.__sql_file_handler.close()

        pass

'''2019.7.8 把excel的数据加载并生成insert sql语句保存到DB中，之后对表的sql在外部做调用执行'''
class DCReconciliationSQL(object):
    # 初始化数据文件到object
    def __init__(self,file_name,ignore_strike_row=True):
        self.__file_name=file_name
        self.excel_handler=openpyxl.load_workbook(file_name, data_only=True)
        self.__ignore_strike_row=ignore_strike_row
        self.__sql_file_handler = open(os.path.join(configure.DOWNLOAD_FOLDER,public_init_script.reconciliation_file_name),'w')
        self.__column_title_name = 'BRR_Status','SN','Module','BRR_CODE','BRR_Desc','BRR_Column1','BRR_Column2','BRR_Column3','Cnt_Column1','Comments','SQL_Target','SQL_Source'
    # 生成校验语句的表，保存excel中定义的校验语句
    def __create_verify_log_table(self):
        # 创建表dc_validation的脚本
        sql = public_init_script.init_dc_reconciliation
    # 取得excel校验语句的数据信息，各个sheet中的格式应该保持统一以方便统一生成
    def get_verify_sheet_data(self, sheetindex = 1 ):
        docconfig = self.excel_handler.worksheets[sheetindex]
        #所有的列名,如果新增加列则直接增加列名,返回的内容只包括列表所在的数据内容
        column_title_name=(self.__column_title_name)

        # 循环所有的行
        config_rows=[]
        for row in docconfig.rows:
            # 循环所有 的列
            cell_value={}
            for cell in row:
                #忽略已经加了删除线的行,只检查前两列是否有删除线，则说明本行是删除的，其它单元格的删除线不识别
                if cell.font.strikethrough and cell.col_idx<=2 and self.__ignore_strike_row:
                    break
                # 排除掉第一行的标题
                if cell.col_idx<=len(column_title_name) and cell.row>=2:
                    title_name=column_title_name[cell.col_idx-1]
                    if cell.value is None:
                        cell_value[title_name]= ''
                    else:
                        cell_value[title_name]=self.__sql_adjust(cell.value)
            # 对于 非空的行则增加到变量里面
            if len(cell_value) > 0:
                config_rows.append(cell_value)
        return config_rows
    # 对SQL注释中的--进行处理，以防止换行符替换后注释扩大，把--替换成/* xxx */的形式, 2019.7.13
    def __replace_remark(self,matched):
        newremark = '/* '+ matched.group() + ' */'
        return newremark
    # 处理sql语句中的特殊字符
    def __sql_adjust(self, sql):
        # 如果sql中有'则做'的处理
        if type(sql)==type('str'):
            newsql = sql.replace("'","''")
            # 替换SQL中的--注释,替换成/* */
            newsql = re.sub(r'--(.*)',self.__replace_remark,newsql)
            #处理excel中有换行的情况，替换成空格
            newsql = newsql.replace("\n"," ").replace('\r',' ')
            # 对于超长的字符串进行换行
            final_sql = ''
            for s in newsql.split(','):
                if len(final_sql+s)>=1000:
                    final_sql = final_sql +'\n' + s+','
                else:
                    final_sql = final_sql + s + ','
            # 移除最后一个,
            final_sql = final_sql[:len(final_sql)-1]
        else:
            final_sql=sql
        return final_sql

        pass
    # 针对已经读取的excel数据，生成insert sql 语句的脚本
    def gen_insert_sql(self):
        all_insert_sql = ''
        # 默认为第二个sheet为校验脚本的sheet, sheet index =1
        validation_rules = self.get_verify_sheet_data(1)
        #所有的列名,如果新增加列则直接增加列名,返回的内容只包括列表所在的数据内容
        column_title_name=(self.__column_title_name)

        for eachvalidation in validation_rules:
            # 对每个列进行处理
            try:
                insert_sql = public_init_script.init_insert_dc_reconciliation.format( \
                        BRR_Status=eachvalidation['BRR_Status '.strip()], \
                        SN=eachvalidation['SN         '.strip()], \
                        Module=eachvalidation['Module     '.strip()], \
                        BRR_CODE=eachvalidation['BRR_CODE   '.strip()], \
                        BRR_Desc=eachvalidation['BRR_Desc   '.strip()], \
                        BRR_Column1=eachvalidation['BRR_Column1'.strip()], \
                        BRR_Column2=eachvalidation['BRR_Column2'.strip()], \
                        BRR_Column3=eachvalidation['BRR_Column3'.strip()], \
                        Cnt_Column1=eachvalidation['Cnt_Column1'.strip()], \
                        Comments=eachvalidation['Comments   '.strip()], \
                        SQL_Target=eachvalidation['SQL_Target '.strip()], \
                        SQL_Source=eachvalidation['SQL_Source '.strip()], \
                    )
            except Exception as e:
                print('Data in excel is not matched with format:')
                print(eachvalidation)
            all_insert_sql = all_insert_sql + insert_sql
        all_insert_sql = all_insert_sql + 'commit;\n'
        return all_insert_sql
        pass
    # 生成dc_validation表的创建表语句 及insert 语句
    def get_all_sqls(self):
        all_sql = public_init_script.init_dc_reconciliation + self.gen_insert_sql()
        return all_sql
        pass
    # 2019.7.3 生成脚本文件
    def gen_script_file(self):
        filenameonly = self.__file_name.split('/')[-1]
        self.__sql_file_handler.write("spool {0}.log\n".format(public_init_script.reconciliation_file_name.split('.')[0]))
        # 生成insert sql
        insert_sql = self.get_all_sqls()
        self.__sql_file_handler.write(insert_sql)
        self.__sql_file_handler.write("\nspool off")
        self.__sql_file_handler.write("\n quit; ")
        self.__sql_file_handler.close()

        pass
if __name__=='__main__':

    file_name='./DC_VALIDATION.xlsx'
    # file_name='./投保字段规则合集_20180525.xlsx'

    # veriscript = DCVerifySQL(file_name)
    # sheet_list = [1]
    # data = veriscript.get_all_sqls(sheet_list)
    # print(data)
    # script = veriscript.get_all_sqls()
    # print(script)
    # veriscript.gen_script_file()
    # Test reconciliation
    reconscript = DCReconciliationSQL(file_name='DataMigrationReconciliationReport_V2.5_LS (5).xlsx')
    reconscript.gen_script_file()
    # dcrrreport = DCReport(file_name)
    # dcrrreport.process_report_data()

    # docconfig = DCDocConfigExcel(file_name)
    # docconfig.get_config_data()
    #file_name='sample.xlsx'
    # template1=Template(file_name)
    # mappingcols=template1.get_mapping_cols()

    # for row in mappingcols:
    #     print('%s,%s,%s'%(row.tableName,row.columnName,row.dataType))